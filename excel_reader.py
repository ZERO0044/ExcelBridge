"""
Excel 文件读取模块
支持 .xlsx (openpyxl) 和 .xls (xlrd) 格式
"""

import os
from typing import Dict, List, Any, Optional


def read_excel(filepath: str, skip_hidden: bool = False) -> Dict[str, List[List[Any]]]:
    """
    读取 Excel 文件，返回所有 Sheet 的数据。
    skip_hidden=True 时跳过隐藏行和隐藏列。

    对于 .xlsx：使用 openpyxl（支持隐藏行列检测）。
    对于 .xls：检测文件签名 — 若为 ZIP 格式（实际是 .xlsx 伪装）则用 openpyxl，
    否则回退 xlrd（旧 .xls 格式，隐藏行列检测有局限性）。
    """
    ext = os.path.splitext(filepath)[1].lower()
    if ext == '.xlsx':
        return _read_xlsx(filepath, skip_hidden)
    elif ext == '.xls':
        # 检测真实格式：ZIP 头 = xlsx 伪装
        try:
            with open(filepath, 'rb') as f:
                sig = f.read(4)
            if sig[:2] == b'PK':  # ZIP 魔术字节 → 实际是 xlsx
                # openpyxl 按扩展名拒绝 .xls，需要临时复制为 .xlsx
                import shutil, tempfile
                tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
                tmp.close()
                try:
                    shutil.copy2(filepath, tmp.name)
                    return _read_xlsx(tmp.name, skip_hidden)
                finally:
                    try:
                        os.unlink(tmp.name)
                    except Exception:
                        pass
        except Exception:
            pass
        return _read_xls(filepath, skip_hidden)
    else:
        raise ValueError("不支持的文件格式: {}".format(ext))


def _read_xlsx(filepath: str, skip_hidden: bool = False) -> Dict[str, List[List[Any]]]:
    """用 openpyxl 读取 .xlsx 文件"""
    import openpyxl

    wb = openpyxl.load_workbook(filepath, data_only=True)
    result = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        max_col = ws.max_column or 1

        # 收集隐藏列索引
        hidden_cols = set()
        if skip_hidden:
            for col_letter, col_dim in ws.column_dimensions.items():
                if col_dim.hidden:
                    col_idx = 0
                    for c in col_letter:
                        col_idx = col_idx * 26 + (ord(c.upper()) - ord('A') + 1)
                    hidden_cols.add(col_idx - 1)

        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_col=max_col, values_only=True), start=1):
            # 跳过隐藏行
            if skip_hidden and row_idx in ws.row_dimensions and ws.row_dimensions[row_idx].hidden:
                continue
            row_data = [_normalize_cell(cell) if i not in hidden_cols else ''
                        for i, cell in enumerate(row)]
            rows.append(row_data)

        result[sheet_name] = rows

    wb.close()
    return result


def _read_xls(filepath: str, skip_hidden: bool = False) -> Dict[str, List[List[Any]]]:
    """用 xlrd 读取真正的旧格式 .xls 文件（BIFF）。

    隐藏行列检测采用双层策略：
    1. 优先使用 xlrd 的 rowinfo_map / colinfo_map（对 WPS/Excel 创建的文件有效）
    2. 若为空，回退到直接扫描 BIFF 二进制 ROW/COLINFO 记录
    """
    import xlrd

    # formatting_info=True 才能解析 ROW/COLINFO 记录中的隐藏标志。
    # 始终开启以确保 skip=False/True 两种模式下解析结果一致。
    wb = xlrd.open_workbook(filepath, formatting_info=True)
    result = {}

    for sheet_name in wb.sheet_names():
        ws = wb.sheet_by_name(sheet_name)
        rows = []

        # 收集隐藏行/列
        hidden_rows = set()
        hidden_cols = set()
        if skip_hidden:
            # 第一层：xlrd 的 rowinfo_map / colinfo_map
            try:
                for r_idx, row_info in ws.rowinfo_map.items():
                    if row_info.hidden:
                        hidden_rows.add(r_idx)
            except Exception:
                pass
            try:
                for c_idx, col_info in ws.colinfo_map.items():
                    if col_info.hidden:
                        hidden_cols.add(c_idx)
            except Exception:
                pass

            # 第二层：若 xlrd 没检测到任何隐藏行列，回退到 BIFF 二进制解析
            if not hidden_rows and not hidden_cols:
                _biff_hidden = _parse_biff_hidden(filepath)
                hidden_rows = _biff_hidden['rows']
                hidden_cols = _biff_hidden['cols']

        for row_idx in range(ws.nrows):
            if skip_hidden and row_idx in hidden_rows:
                continue
            row_data = []
            for col_idx in range(ws.ncols):
                cell = ws.cell(row_idx, col_idx)
                if skip_hidden and col_idx in hidden_cols:
                    row_data.append('')
                else:
                    row_data.append(_normalize_xlrd_cell(cell))
            rows.append(row_data)

        result[sheet_name] = rows

    return result


def _parse_biff_hidden(filepath: str) -> dict:
    """回退方案：直接扫描 BIFF 二进制文件中的 ROW 和 COLINFO 记录。

    BIFF8 ROW 记录 (0x0208, 16字节):
        offset 0:  rw (2B, row index 0-based)
        offset 12: flags (2B, bit 0x0020 = hidden)

    BIFF8 COLINFO 记录 (0x007D, 12字节):
        offset 0:  colFirst (2B)
        offset 2:  colLast (2B)
        offset 10: flags (2B, bit 0x0001 = hidden)

    返回 {'rows': set(), 'cols': set()}
    """
    import struct
    result = {'rows': set(), 'cols': set()}
    try:
        with open(filepath, 'rb') as f:
            # 跳过 OLE header (512 bytes) — BIFF 数据从第一个 sector 开始
            data = f.read()
    except Exception:
        return result

    # 扫描 ROW 记录 (0x0208)
    idx = 0
    while True:
        pos = data.find(b'\x08\x02', idx)  # 0x0208 little-endian
        if pos == -1:
            break
        if pos + 18 <= len(data):
            rec_len = struct.unpack_from('<H', data, pos + 2)[0]
            if rec_len == 16:  # ROW 记录固定长度
                rw = struct.unpack_from('<H', data, pos + 4)[0]
                flags = struct.unpack_from('<H', data, pos + 14)[0]
                if flags & 0x0020:
                    result['rows'].add(rw)
        idx = pos + 1

    # 扫描 COLINFO 记录 (0x007D)
    idx = 0
    while True:
        pos = data.find(b'\x7D\x00', idx)  # 0x007D little-endian
        if pos == -1:
            break
        if pos + 18 <= len(data):
            rec_len = struct.unpack_from('<H', data, pos + 2)[0]
            if rec_len == 12:  # COLINFO 记录固定长度
                col_first = struct.unpack_from('<H', data, pos + 4)[0]
                col_last = struct.unpack_from('<H', data, pos + 6)[0]
                flags = struct.unpack_from('<H', data, pos + 14)[0]
                if flags & 0x0001:
                    for c in range(col_first, col_last + 1):
                        result['cols'].add(c)
        idx = pos + 1

    return result


def _normalize_cell(value: Any) -> str:
    """将 openpyxl 单元格值转为字符串（空值返回空字符串）"""
    if value is None:
        return ''
    if isinstance(value, (int, float)):
        # 整数不显示小数点
        if isinstance(value, float) and value == int(value):
            return str(int(value))
        return str(value)
    return str(value).strip()


def _normalize_xlrd_cell(cell: Any) -> str:
    """将 xlrd 单元格值转为字符串"""
    import xlrd

    if cell.ctype == xlrd.XL_CELL_EMPTY:
        return ''
    elif cell.ctype == xlrd.XL_CELL_DATE:
        import datetime
        try:
            dt = xlrd.xldate_as_datetime(cell.value, 0)
            return dt.strftime('%Y-%m-%d')
        except Exception:
            return str(cell.value)
    elif cell.ctype == xlrd.XL_CELL_NUMBER:
        if cell.value == int(cell.value):
            return str(int(cell.value))
        return str(cell.value)
    else:
        return str(cell.value).strip()


def find_files(root_dir: str, recursive: bool = True) -> List[str]:
    """
    在目录下查找所有 .xlsx 和 .xls 文件

    Args:
        root_dir: 根目录路径
        recursive: 是否递归搜索子目录

    Returns:
        文件路径列表（按相对路径排序）
    """
    root_dir = os.path.abspath(root_dir)
    extensions = {'.xlsx', '.xls'}
    files = []

    if recursive:
        for dirpath, _dirnames, filenames in os.walk(root_dir):
            for fname in filenames:
                # 跳过临时文件（WPS 锁定文件等）
                if fname.startswith('.~') or fname.startswith('~$'):
                    continue
                ext = os.path.splitext(fname)[1].lower()
                if ext in extensions:
                    full_path = os.path.join(dirpath, fname)
                    files.append(full_path)
    else:
        for fname in os.listdir(root_dir):
            full_path = os.path.join(root_dir, fname)
            if os.path.isfile(full_path):
                if fname.startswith('.~') or fname.startswith('~$'):
                    continue
                ext = os.path.splitext(fname)[1].lower()
                if ext in extensions:
                    files.append(full_path)

    files.sort()
    return files


def get_relative_path(filepath: str, base_dir: str) -> str:
    """获取文件相对于 base_dir 的相对路径"""
    try:
        return os.path.relpath(filepath, base_dir)
    except ValueError:
        return os.path.basename(filepath)
