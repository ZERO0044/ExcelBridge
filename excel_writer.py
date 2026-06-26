"""
Excel 写入引擎
按匹配键查找目标行并写入数据，保留模板样式
"""

import os
import datetime
from typing import List, Dict, Optional, Callable

from rule_engine import ColumnMapping, MatchEntry


def process_and_write(
    target_path: str,
    mapping: ColumnMapping,
    entries: List[MatchEntry],
    target_sheet: str = None,
    progress_callback: Optional[Callable[[str], None]] = None,
) -> str:
    """
    执行批量写入操作。

    Args:
        target_path: 目标模板文件路径
        mapping: 列映射规则
        entries: 用户确认的要写入的匹配条目列表（selected=True 的才会写入）
        target_sheet: 目标 Sheet 名（默认第一个 Sheet）
        progress_callback: 进度回调函数 (message) -> None

    Returns:
        输出文件路径
    """
    import openpyxl

    if progress_callback is None:
        progress_callback = lambda msg: None

    progress_callback("正在打开目标模板...")
    wb = openpyxl.load_workbook(target_path)

    # 确定目标 Sheet
    if target_sheet is None or target_sheet not in wb.sheetnames:
        target_sheet = wb.sheetnames[0]

    ws = wb[target_sheet]
    progress_callback("目标 Sheet: {}".format(target_sheet))

    # 构建目标索引：键 → 行号 (1-based)
    progress_callback("正在构建目标索引...")
    target_index: Dict[str, int] = {}

    for row_idx in range(1, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=mapping.target_key_col + 1)
        key = _normalize_cell(cell.value)
        if key:
            target_index[key] = row_idx

    progress_callback("目标中有 {} 条索引记录".format(len(target_index)))

    # 写入数据
    written = 0
    skipped = 0
    not_found = 0
    total = sum(1 for e in entries if e.selected and e.has_data)

    for entry in entries:
        if not entry.selected or not entry.has_data:
            if entry.selected and not entry.has_data:
                skipped += 1
            continue

        target_row = target_index.get(entry.key, 0)
        if target_row == 0:
            not_found += 1
            progress_callback("[未找到] {} 在目标中不存在".format(entry.key))
            continue

        # 写入值到目标单元格
        dest_cell = ws.cell(row=target_row, column=mapping.target_dest_col + 1)
        dest_cell.value = entry.value
        written += 1

        progress_callback("[写入] {} → 目标行{}: \"{}\"".format(
            entry.key, target_row, _truncate(entry.value, 30)))

    progress_callback("写入完成: {} 条, 跳过 {} 条, 未找到 {} 条".format(
        written, skipped, not_found))

    # 生成输出文件名
    base, ext = os.path.splitext(os.path.basename(target_path))
    timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    output_dir = os.path.dirname(os.path.abspath(target_path))
    output_path = os.path.join(output_dir, "{}_结果_{}{}".format(base, timestamp, ext))

    progress_callback("正在保存: {}".format(output_path))
    wb.save(output_path)
    wb.close()

    return output_path


def _normalize_cell(value):
    """规范化单元格值，与 excel_reader 保持一致，确保键匹配不因类型差异失败"""
    if value is None:
        return ''
    if isinstance(value, (int, float)):
        if isinstance(value, float) and value == int(value):
            return str(int(value))
        return str(value)
    return str(value).strip()


def _truncate(text: str, max_len: int) -> str:
    """截断字符串"""
    if len(text) <= max_len:
        return text
    return text[:max_len - 3] + '...'
